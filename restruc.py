import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
import os

# arquivo exportado do Forms
arquivo = "Dados do Responsável do Setor.xlsx"

# ler planilha
df = pd.read_excel(arquivo)

# MOSTRAR COLUNAS PARA CONFIRMAR (remova depois se quiser)
print("Colunas encontradas:")
for col in df.columns:
    print(f"  - '{col}'")

# Criar novo workbook
wb = Workbook()
ws = wb.active
ws.title = "Cadastro de Rede"

# ========== ESTILOS COM CALIBRI ==========
# Cores
cinza_claro = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Cinza claro
cinza_escuro = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Cinza mais escuro
branco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# Fontes CALIBRI - TAMANHOS CORRIGIDOS
titulo_font = Font(name='Calibri', size=24, bold=True)  # TAMANHO 24 para CADASTRO DE REDE
subtitulo_font = Font(name='Calibri', size=11, bold=True)  # TAMANHO 11 para subtítulos
bold_font = Font(name='Calibri', size=11, bold=True)  # TAMANHO 11 para negritos
normal_font = Font(name='Calibri', size=11)  # TAMANHO 11 para corpo

# Bordas COMPLETAS (todas as laterais)
borda_completa = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ========== INSERIR LOGO DA PREFEITURA ==========
# Verificar se o arquivo da logo existe
caminho_logo = "logo_prefeitura.png"  # Coloque o arquivo na mesma pasta
if os.path.exists(caminho_logo):
    img = Image(caminho_logo)
    img.width = 80  # Ajuste o tamanho conforme necessário
    img.height = 80
    ws.add_image(img, 'A1')  # Posiciona na célula A1
    print("✅ Logo inserida com sucesso!")
else:
    print("⚠️ Arquivo da logo não encontrado. Coloque 'logo_prefeitura.png' na pasta do projeto.")

linha = 1

# Agrupar por responsável
responsaveis = df['Nome do responsável'].unique()

for responsavel in responsaveis:
    dados = df[df['Nome do responsável'] == responsavel]
    primeiro = dados.iloc[0]
    
    # ========== LINHA DO TÍTULO COM CINZA CLARO ==========
    # Mesclar células para o título (A até E)
    ws.merge_cells(f'A{linha}:E{linha}')
    
    # Configurar célula do título
    celula_titulo = ws.cell(linha, 1, "CADASTRO DE REDE")
    celula_titulo.font = titulo_font  # Calibri Bold 24
    celula_titulo.alignment = Alignment(horizontal='center', vertical='center')
    celula_titulo.fill = cinza_claro
    celula_titulo.border = borda_completa
    
    # Aumentar altura da linha do título
    ws.row_dimensions[linha].height = 40  # Altura maior para caber fonte 24
    
    linha += 2  # Espaço após o título
    
    # ========== DADOS DO RESPONSÁVEL ==========
    # Célula de título da seção (com cinza mais escuro)
    ws.merge_cells(f'A{linha}:E{linha}')
    celula_subtitulo = ws.cell(linha, 1, "DADOS DO RESPONSÁVEL DO SETOR")
    celula_subtitulo.font = subtitulo_font  # Calibri Bold 11
    celula_subtitulo.alignment = Alignment(horizontal='center')
    celula_subtitulo.fill = cinza_escuro  # Cinza mais escuro
    celula_subtitulo.border = borda_completa
    linha += 1
    
    # Dados do responsável (formato chave-valor com bordas)
    dados_resp = [
        ("Nome:", primeiro['Nome do responsável']),
        ("Matrícula:", primeiro['Matrícula do responsável']),
        ("Login de Rede:", primeiro['Login de rede do responsável']),
        ("E-mail institucional:", primeiro['E-mail institucional do responsável']),
        ("Departamento:", primeiro['Departamento']),
        ("Quantidade de servidores:", len(dados))
    ]
    
    for campo, valor in dados_resp:
        # Célula do campo (coluna A) - Calibri Bold 11
        celula_campo = ws.cell(linha, 1, campo)
        celula_campo.font = bold_font  # Calibri Bold 11
        celula_campo.border = borda_completa
        celula_campo.fill = branco
        
        # Célula do valor (coluna B) - Calibri Normal 11
        celula_valor = ws.cell(linha, 2, valor)
        celula_valor.font = normal_font  # Calibri Normal 11
        celula_valor.border = borda_completa
        celula_valor.fill = branco
        
        # Bordas nas colunas C, D, E (vazias, mas com borda)
        for col in range(3, 6):
            celula_vazia = ws.cell(linha, col, "")
            celula_vazia.border = borda_completa
            celula_vazia.fill = branco
        
        linha += 1
    
    linha += 1  # Espaço após dados do responsável
    
    # ========== TABELA DE SERVIDORES ==========
    # Célula de título da seção (com cinza mais escuro)
    ws.merge_cells(f'A{linha}:E{linha}')
    celula_subtitulo = ws.cell(linha, 1, "DADOS DOS SERVIDORES")
    celula_subtitulo.font = subtitulo_font  # Calibri Bold 11
    celula_subtitulo.alignment = Alignment(horizontal='center')
    celula_subtitulo.fill = cinza_escuro  # Cinza mais escuro
    celula_subtitulo.border = borda_completa
    linha += 1
    
    # Cabeçalho da tabela - Calibri Bold 11
    cabecalhos = ["Nome", "Matrícula", "Login de Rede", "E-mail Institucional", "Coordenação"]
    for col, cab in enumerate(cabecalhos, 1):
        celula = ws.cell(linha, col, cab)
        celula.font = bold_font  # Calibri Bold 11
        celula.alignment = Alignment(horizontal='center')
        celula.fill = cinza_claro
        celula.border = borda_completa
    
    linha += 1
    
    # Dados dos servidores (todas as linhas com bordas) - Calibri Normal 11
    for _, row in dados.iterrows():
        # Nome
        celula = ws.cell(linha, 1, row['Nome do servidor'])
        celula.font = normal_font  # Calibri Normal 11
        celula.border = borda_completa
        celula.fill = branco
        
        # Matrícula
        celula = ws.cell(linha, 2, row['Matrícula do servidor'])
        celula.font = normal_font  # Calibri Normal 11
        celula.border = borda_completa
        celula.fill = branco
        
        # Login de Rede
        celula = ws.cell(linha, 3, row['Login de rede'])
        celula.font = normal_font  # Calibri Normal 11
        celula.border = borda_completa
        celula.fill = branco
        
        # E-mail
        celula = ws.cell(linha, 4, row['E-mail institucional'])
        celula.font = normal_font  # Calibri Normal 11
        celula.border = borda_completa
        celula.fill = branco
        
        # Coordenação
        celula = ws.cell(linha, 5, row['Coordenação do servidor'])
        celula.font = normal_font  # Calibri Normal 11
        celula.border = borda_completa
        celula.fill = branco
        
        linha += 1
    
    # Espaço entre grupos de responsáveis
    linha += 3

# ========== AJUSTES FINAIS ==========
# Ajustar largura das colunas
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 45
ws.column_dimensions['E'].width = 25

# Congelar painel para facilitar visualização
ws.freeze_panes = 'A2'

# Salvar
wb.save("cadastro_rede_tamanhos_corretos.xlsx")
print("\n✅ Arquivo criado: cadastro_rede_tamanhos_corretos.xlsx")
print("📌 Fonte: Calibri")
print("   - CADASTRO DE REDE: 24")
print("   - Demais textos: 11")
print("🎨 Design aplicado com cores e bordas")